#!/usr/bin/env python3
"""
EP28-A3c 스타일 참고구간(reference interval) 분석 스크립트

목적
- 엑셀 파일에 수집된 참고집단 검사결과를 읽어 partition별 참고구간을 계산합니다.
- 기본 방식은 CLSI EP28-A3c에서 널리 쓰는 비모수 percentile 방식입니다.
- 양측 95% 참고구간은 2.5th~97.5th percentile로 계산합니다.
- 각 reference limit의 90% confidence interval은 binomial/order-statistic 방식으로 계산합니다.
- 이상치는 자동 삭제하지 않고 flag만 표시하는 것을 기본값으로 합니다.
- partition 컬럼이 주어지면 Harris & Boyd approach로 partition 간 분리 필요성을 pairwise 평가합니다.

필요 패키지
    pip install pandas numpy scipy openpyxl

기본 사용 예
    python ep28_reference_interval_analysis.py data.xlsx --value-col Result

성별/연령군별 partition 분석 예
    python ep28_reference_interval_analysis.py data.xlsx \
        --sheet Sheet1 \
        --value-col Result \
        --partition-cols Sex AgeGroup \
        --units U/L \
        --out ep28_RI_report.xlsx

수동 제외 컬럼이 있는 경우 예
    python ep28_reference_interval_analysis.py data.xlsx \
        --value-col Result \
        --partition-cols Sex \
        --manual-exclude-col Exclude

Harris & Boyd partition 평가에 log 변환을 적용하는 예
    python ep28_reference_interval_analysis.py data.xlsx \
        --value-col Result \
        --partition-cols Sex \
        --hb-transform log

주의
- EP28-A3c 원칙상 partition당 n>=120일 때 비모수 참고구간 설정이 가장 안정적입니다.
- n<120 결과도 계산은 하지만, 보고서에는 '주의/예비적 해석'으로 표시합니다.
- 이상치 flag는 검토용입니다. --exclude-flagged-outliers 옵션은 검체/분석 오류 등 배제 근거가 확인된 후에만 사용하세요.
- Harris & Boyd approach는 하위집단 값이 대략 Gaussian 분포라는 전제가 있으므로, 심한 왜도에서는 log/Box-Cox 변환 또는 다른 partition 평가법도 함께 검토하세요.
"""

from __future__ import annotations

import argparse
import math
from itertools import combinations
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from scipy import stats
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


TRUE_STRINGS = {"1", "true", "t", "yes", "y", "exclude", "excluded", "제외", "예", "yes"}


@dataclass
class AnalysisConfig:
    value_col: str
    partition_cols: list[str]
    manual_exclude_col: Optional[str]
    central_fraction: float = 0.95
    limit_ci: float = 0.90
    min_n_nonparametric: int = 120
    outlier_method: str = "both"  # none, tukey, reed, both
    exclude_flagged_outliers: bool = False
    units: str = ""
    decimal_places: int = 3
    harris_boyd_enabled: bool = True
    harris_boyd_transform: str = "none"  # none, log, boxcox
    harris_boyd_min_n: int = 3
    normality_alpha: float = 0.05


# -----------------------------
# Utility functions
# -----------------------------

def parse_sheet_arg(sheet: str):
    """pandas read_excel에 넣을 sheet_name을 문자열/정수로 변환합니다."""
    if sheet.isdigit():
        return int(sheet)
    return sheet


def as_bool_exclude(x) -> bool:
    """수동 제외 컬럼 값을 bool로 해석합니다."""
    if pd.isna(x):
        return False
    if isinstance(x, (bool, np.bool_)):
        return bool(x)
    if isinstance(x, (int, float, np.integer, np.floating)):
        return bool(x)
    return str(x).strip().lower() in TRUE_STRINGS


def make_partition_key(row: pd.Series, partition_cols: list[str]) -> str:
    if not partition_cols:
        return "ALL"
    parts = []
    for col in partition_cols:
        val = row[col]
        if pd.isna(val):
            val = "<missing>"
        parts.append(f"{col}={val}")
    return " | ".join(parts)


def percentile_nplus1(sorted_values: np.ndarray, p: float) -> float:
    """
    EP28-A3c 실무에서 흔히 쓰는 p*(n+1) rank 기반 percentile 추정.
    rank가 정수가 아니면 선형 보간합니다.

    p: 0~1 사이 percentile. 예: 0.025, 0.975
    """
    x = np.asarray(sorted_values, dtype=float)
    n = len(x)
    if n == 0:
        return np.nan
    if n == 1:
        return float(x[0])

    h = p * (n + 1)
    if h <= 1:
        return float(x[0])
    if h >= n:
        return float(x[-1])

    lower_rank = int(math.floor(h))  # 1-based rank
    fraction = h - lower_rank

    lower_value = x[lower_rank - 1]
    upper_value = x[lower_rank]
    return float(lower_value + fraction * (upper_value - lower_value))


def order_stat_percentile_ci(sorted_values: np.ndarray, p: float, confidence: float = 0.90) -> dict:
    """
    Binomial/order-statistic 방식으로 percentile reference limit의 confidence interval 계산.

    반환값의 rank는 1-based rank입니다.
    예: n=120, p=0.025, confidence=0.90이면 보통 1st~7th order statistic 범위가 나옵니다.
    """
    x = np.asarray(sorted_values, dtype=float)
    n = len(x)
    if n == 0:
        return {
            "ci_low": np.nan,
            "ci_high": np.nan,
            "rank_low": np.nan,
            "rank_high": np.nan,
            "actual_coverage": np.nan,
        }

    alpha = 1.0 - confidence

    # j, k는 이론상 P[X_(j) <= q_p <= X_(k)] 계산에 쓰이는 order-statistic rank입니다.
    # scipy binom.ppf는 'p보다 작거나 같은 관측 수'의 분위수를 반환합니다.
    j = int(stats.binom.ppf(alpha / 2.0, n, p))
    k = int(stats.binom.ppf(1.0 - alpha / 2.0, n, p)) + 1

    # sample 밖으로 나가는 경우에는 가장 가까운 order statistic으로 clamp합니다.
    rank_low = max(1, min(n, j))
    rank_high = max(1, min(n, k))
    if rank_low > rank_high:
        rank_low, rank_high = rank_high, rank_low

    # 이론적 coverage: sum_{i=j}^{k-1} Bin(n,p). j/k clamp 전 기준으로 계산하되 범위 제한.
    j_cov = max(0, min(n, j))
    k_cov = max(0, min(n + 1, k))
    if k_cov <= j_cov:
        actual_coverage = np.nan
    else:
        actual_coverage = float(sum(stats.binom.pmf(i, n, p) for i in range(j_cov, k_cov)))

    return {
        "ci_low": float(x[rank_low - 1]),
        "ci_high": float(x[rank_high - 1]),
        "rank_low": int(rank_low),
        "rank_high": int(rank_high),
        "actual_coverage": actual_coverage,
    }


def format_number(x, digits: int) -> str:
    if pd.isna(x):
        return ""
    return f"{float(x):.{digits}f}"


# -----------------------------
# Harris & Boyd partition assessment
# -----------------------------

def shapiro_p_value(values: np.ndarray) -> tuple[float, str]:
    """
    Shapiro-Wilk 정규성 검정 p-value를 반환합니다.

    Harris & Boyd 접근은 원칙적으로 각 하위집단 값이 대략 Gaussian 분포라는
    전제를 가지므로, 해석 보조용으로만 기록합니다.
    """
    x = np.asarray(values, dtype=float)
    x = x[np.isfinite(x)]
    n = len(x)
    if n < 3:
        return np.nan, "not tested: n<3"
    if n > 5000:
        return np.nan, "not tested: n>5000; use histogram/Q-Q plot or separate normality assessment"
    if np.nanstd(x, ddof=1) == 0:
        return np.nan, "not tested: zero variance"
    try:
        return float(stats.shapiro(x).pvalue), "tested"
    except Exception as exc:  # pragma: no cover - diagnostic only
        return np.nan, f"not tested: {exc}"


def transform_pair_for_harris_boyd(
    values1: np.ndarray,
    values2: np.ndarray,
    method: str,
) -> tuple[np.ndarray, np.ndarray, float, str]:
    """
    Harris & Boyd 계산 전 두 그룹을 동일한 방식으로 변환합니다.

    Parameters
    ----------
    values1, values2:
        비교할 두 partition의 원자료 값.
    method:
        'none', 'log', 'boxcox' 중 하나.

    Returns
    -------
    transformed1, transformed2, lambda_value, note
        lambda_value는 Box-Cox 변환에서만 의미가 있으며, 그 외에는 NaN입니다.
    """
    x1 = np.asarray(values1, dtype=float)
    x2 = np.asarray(values2, dtype=float)
    x1 = x1[np.isfinite(x1)]
    x2 = x2[np.isfinite(x2)]

    if method == "none":
        return x1, x2, np.nan, "no transformation"

    pooled = np.concatenate([x1, x2])
    if np.any(pooled <= 0):
        return x1, x2, np.nan, (
            f"{method} transformation skipped because one or more values are <=0; "
            "Harris-Boyd statistics shown on original scale"
        )

    if method == "log":
        return np.log(x1), np.log(x2), np.nan, "natural-log transformed"

    if method == "boxcox":
        try:
            transformed, lambda_value = stats.boxcox(pooled)
            t1 = transformed[: len(x1)]
            t2 = transformed[len(x1) :]
            return t1, t2, float(lambda_value), "Box-Cox transformed using pooled pair data"
        except Exception as exc:  # pragma: no cover - diagnostic only
            return x1, x2, np.nan, (
                f"Box-Cox transformation failed ({exc}); "
                "Harris-Boyd statistics shown on original scale"
            )

    raise ValueError(f"Unsupported Harris-Boyd transform: {method}")


def harris_boyd_pairwise_test(
    partition_1: str,
    values1: np.ndarray,
    partition_2: str,
    values2: np.ndarray,
    cfg: AnalysisConfig,
) -> dict:
    """
    두 partition 사이의 Harris & Boyd pairwise 평가를 수행합니다.

    Z_calculated = |mean1 - mean2| / sqrt(sd1^2/n1 + sd2^2/n2)
    Z_critical   = 3 * sqrt((n1 + n2) / 240)

    기본 판정은 Z_calculated > Z_critical이면 별도 RI 분리를 권고합니다.
    SD ratio >= 1.5 및 s_large/(s_large-s_small) < 3은 보조 flag로 표시합니다.
    """
    raw1 = np.asarray(values1, dtype=float)
    raw2 = np.asarray(values2, dtype=float)
    raw1 = raw1[np.isfinite(raw1)]
    raw2 = raw2[np.isfinite(raw2)]
    n1 = len(raw1)
    n2 = len(raw2)

    base = {
        "partition_1": partition_1,
        "partition_2": partition_2,
        "n1": n1,
        "n2": n2,
        "transform": cfg.harris_boyd_transform,
        "boxcox_lambda": np.nan,
        "original_mean1": float(np.mean(raw1)) if n1 else np.nan,
        "original_mean2": float(np.mean(raw2)) if n2 else np.nan,
        "original_sd1": float(np.std(raw1, ddof=1)) if n1 >= 2 else np.nan,
        "original_sd2": float(np.std(raw2, ddof=1)) if n2 >= 2 else np.nan,
        "original_median1": float(np.median(raw1)) if n1 else np.nan,
        "original_median2": float(np.median(raw2)) if n2 else np.nan,
    }

    if n1 < cfg.harris_boyd_min_n or n2 < cfg.harris_boyd_min_n:
        base.update({
            "analysis_mean1": np.nan,
            "analysis_mean2": np.nan,
            "analysis_sd1": np.nan,
            "analysis_sd2": np.nan,
            "z_calculated": np.nan,
            "z_critical": np.nan,
            "z_gt_zcritical": False,
            "sd_ratio_large_to_small": np.nan,
            "sd_ratio_ge_1_5": False,
            "sd_secondary_metric": np.nan,
            "sd_secondary_lt_3": False,
            "normality_p1": np.nan,
            "normality_p2": np.nan,
            "normality_comment": "not tested: insufficient n",
            "decision": "Insufficient data for Harris-Boyd comparison",
            "note": f"Each group needs at least {cfg.harris_boyd_min_n} numeric values.",
        })
        return base

    t1, t2, lambda_value, transform_note = transform_pair_for_harris_boyd(
        raw1, raw2, cfg.harris_boyd_transform
    )
    base["boxcox_lambda"] = lambda_value

    mean1 = float(np.mean(t1))
    mean2 = float(np.mean(t2))
    sd1 = float(np.std(t1, ddof=1))
    sd2 = float(np.std(t2, ddof=1))

    p1, p1_note = shapiro_p_value(t1)
    p2, p2_note = shapiro_p_value(t2)
    normality_comment = f"group1: {p1_note}; group2: {p2_note}"
    if np.isfinite(p1) and np.isfinite(p2):
        if p1 < cfg.normality_alpha or p2 < cfg.normality_alpha:
            normality_comment += (
                f"; caution: Shapiro p<{cfg.normality_alpha:g} in at least one group"
            )
        else:
            normality_comment += "; no Shapiro-Wilk evidence against normality"

    se = math.sqrt((sd1 ** 2) / n1 + (sd2 ** 2) / n2) if n1 and n2 else np.nan
    if se == 0 or not np.isfinite(se):
        z_calc = np.nan
    else:
        z_calc = abs(mean1 - mean2) / se
    z_crit = 3.0 * math.sqrt((n1 + n2) / 240.0)
    z_gt = bool(np.isfinite(z_calc) and z_calc > z_crit)

    sd_small = min(sd1, sd2)
    sd_large = max(sd1, sd2)
    if sd_small > 0 and np.isfinite(sd_small) and np.isfinite(sd_large):
        sd_ratio = sd_large / sd_small
    else:
        sd_ratio = np.nan
    sd_ratio_flag = bool(np.isfinite(sd_ratio) and sd_ratio >= 1.5)

    if sd_large > sd_small and np.isfinite(sd_large) and np.isfinite(sd_small):
        sd_secondary = sd_large / (sd_large - sd_small)
    else:
        sd_secondary = np.inf
    sd_secondary_flag = bool(np.isfinite(sd_secondary) and sd_secondary < 3.0)

    if z_gt:
        decision = "Partition recommended: Z_calculated > Z_critical"
    elif sd_ratio_flag or sd_secondary_flag:
        decision = "Review partitioning: SD difference is large despite Z criterion not met"
    else:
        decision = "No partition recommended by Harris-Boyd criteria"

    note_parts = [transform_note]
    if cfg.harris_boyd_transform == "none":
        note_parts.append("Harris-Boyd assumes approximate Gaussian distributions; consider log/Box-Cox if skewed.")
    if sd_ratio_flag:
        note_parts.append("Large SD ratio flag: larger SD is >=1.5 times smaller SD.")
    if sd_secondary_flag:
        note_parts.append("Secondary SD flag: s_large/(s_large-s_small) < 3.")

    base.update({
        "analysis_mean1": mean1,
        "analysis_mean2": mean2,
        "analysis_sd1": sd1,
        "analysis_sd2": sd2,
        "z_calculated": float(z_calc) if np.isfinite(z_calc) else np.nan,
        "z_critical": float(z_crit),
        "z_gt_zcritical": z_gt,
        "sd_ratio_large_to_small": float(sd_ratio) if np.isfinite(sd_ratio) else np.nan,
        "sd_ratio_ge_1_5": sd_ratio_flag,
        "sd_secondary_metric": float(sd_secondary) if np.isfinite(sd_secondary) else np.nan,
        "sd_secondary_lt_3": sd_secondary_flag,
        "normality_p1": p1,
        "normality_p2": p2,
        "normality_comment": normality_comment,
        "decision": decision,
        "note": " ".join(note_parts),
    })
    return base


def prepare_analysis_values_by_partition(flagged_df: pd.DataFrame, cfg: AnalysisConfig) -> dict[str, np.ndarray]:
    """
    RI 계산에 사용된 것과 동일한 규칙으로 partition별 분석값을 추출합니다.
    """
    eligible_mask = (~flagged_df["_missing_or_non_numeric"]) & (~flagged_df["_manual_exclude"])
    if cfg.exclude_flagged_outliers:
        eligible_mask = eligible_mask & (flagged_df["_outlier_flag"] == "")

    out: dict[str, np.ndarray] = {}
    for partition_name, part_df in flagged_df.loc[eligible_mask].groupby("_partition", dropna=False, sort=True):
        values = part_df["_value_numeric"].astype(float).to_numpy()
        values = values[np.isfinite(values)]
        out[str(partition_name)] = values
    return out


def assess_harris_boyd_partitions(flagged_df: pd.DataFrame, cfg: AnalysisConfig) -> pd.DataFrame:
    """
    partition이 2개 이상 있을 때 모든 pairwise 조합에 대해 Harris & Boyd 평가를 수행합니다.
    """
    if not cfg.partition_cols:
        return pd.DataFrame([{
            "decision": "Not performed: no partition columns were specified",
            "note": "Use --partition-cols to enable Harris-Boyd partition assessment.",
        }])
    if not cfg.harris_boyd_enabled:
        return pd.DataFrame([{
            "decision": "Not performed: Harris-Boyd assessment disabled",
            "note": "Remove --skip-harris-boyd to enable this assessment.",
        }])

    values_by_partition = prepare_analysis_values_by_partition(flagged_df, cfg)
    partition_names = sorted(values_by_partition.keys())
    if len(partition_names) < 2:
        return pd.DataFrame([{
            "decision": "Not performed: fewer than two analyzable partitions",
            "note": "At least two partitions with numeric values are required.",
        }])

    rows = []
    for p1, p2 in combinations(partition_names, 2):
        rows.append(harris_boyd_pairwise_test(
            p1, values_by_partition[p1], p2, values_by_partition[p2], cfg
        ))
    return pd.DataFrame(rows)


# -----------------------------
# Outlier flagging
# -----------------------------

def tukey_outlier_flags(values: pd.Series) -> pd.Series:
    """
    Tukey fence 방식 이상치 flag.
    - mild: Q1-1.5*IQR 미만 또는 Q3+1.5*IQR 초과
    - extreme: Q1-3*IQR 미만 또는 Q3+3*IQR 초과
    """
    s = values.astype(float)
    flags = pd.Series("", index=s.index, dtype="object")
    if s.dropna().shape[0] < 4:
        return flags

    q1 = s.quantile(0.25)
    q3 = s.quantile(0.75)
    iqr = q3 - q1
    if iqr == 0 or pd.isna(iqr):
        return flags

    mild_low = q1 - 1.5 * iqr
    mild_high = q3 + 1.5 * iqr
    extreme_low = q1 - 3.0 * iqr
    extreme_high = q3 + 3.0 * iqr

    mild_mask = (s < mild_low) | (s > mild_high)
    extreme_mask = (s < extreme_low) | (s > extreme_high)

    flags.loc[mild_mask] = "Tukey_mild"
    flags.loc[extreme_mask] = "Tukey_extreme"
    return flags


def reed_dixon_outlier_flags(values: pd.Series) -> pd.Series:
    """
    Reed/Dixon 스타일의 간단한 range-ratio flag.
    - 가장 작은 값과 두 번째 값 차이 / 전체 범위 > 1/3이면 low-end outlier 후보
    - 가장 큰 값과 두 번째 큰 값 차이 / 전체 범위 > 1/3이면 high-end outlier 후보

    이 함수는 '후보 flag'만 만들며 자동 제외하지 않습니다.
    """
    s = values.astype(float).dropna().sort_values()
    flags = pd.Series("", index=values.index, dtype="object")
    if len(s) < 3:
        return flags

    x_min = s.iloc[0]
    x_second = s.iloc[1]
    x_penultimate = s.iloc[-2]
    x_max = s.iloc[-1]
    data_range = x_max - x_min
    if data_range == 0 or pd.isna(data_range):
        return flags

    low_ratio = (x_second - x_min) / data_range
    high_ratio = (x_max - x_penultimate) / data_range

    if low_ratio > (1.0 / 3.0):
        flags.loc[s.index[0]] = "Reed_low"
    if high_ratio > (1.0 / 3.0):
        flags.loc[s.index[-1]] = "Reed_high"

    return flags


def combine_outlier_flags(values: pd.Series, method: str) -> pd.Series:
    if method == "none":
        return pd.Series("", index=values.index, dtype="object")

    components = []
    if method in {"tukey", "both"}:
        components.append(tukey_outlier_flags(values))
    if method in {"reed", "both"}:
        components.append(reed_dixon_outlier_flags(values))

    if not components:
        return pd.Series("", index=values.index, dtype="object")

    combined = pd.Series("", index=values.index, dtype="object")
    for flag_series in components:
        for idx, flag in flag_series.items():
            if flag:
                if combined.loc[idx]:
                    combined.loc[idx] = combined.loc[idx] + ";" + flag
                else:
                    combined.loc[idx] = flag
    return combined


# -----------------------------
# Main analysis
# -----------------------------

def analyze_one_partition(data: pd.DataFrame, cfg: AnalysisConfig, partition_name: str) -> tuple[dict, pd.DataFrame]:
    """단일 partition의 RI와 row-level flag를 계산합니다."""
    df = data.copy()

    # 숫자 변환
    df["_value_numeric"] = pd.to_numeric(df[cfg.value_col], errors="coerce")
    df["_missing_or_non_numeric"] = df["_value_numeric"].isna()

    # 수동 제외
    if cfg.manual_exclude_col:
        df["_manual_exclude"] = df[cfg.manual_exclude_col].apply(as_bool_exclude)
    else:
        df["_manual_exclude"] = False

    eligible_mask = (~df["_missing_or_non_numeric"]) & (~df["_manual_exclude"])
    eligible_values = df.loc[eligible_mask, "_value_numeric"]

    # 이상치 flag: eligible 값에 대해서만 계산
    df["_outlier_flag"] = ""
    if len(eligible_values) > 0:
        outlier_flags = combine_outlier_flags(eligible_values, cfg.outlier_method)
        df.loc[outlier_flags.index, "_outlier_flag"] = outlier_flags

    if cfg.exclude_flagged_outliers:
        analysis_mask = eligible_mask & (df["_outlier_flag"] == "")
    else:
        analysis_mask = eligible_mask

    analysis_values = df.loc[analysis_mask, "_value_numeric"].astype(float).sort_values().to_numpy()
    n_total = len(df)
    n_missing = int(df["_missing_or_non_numeric"].sum())
    n_manual_excluded = int(df["_manual_exclude"].sum())
    n_outlier_flagged = int((df["_outlier_flag"] != "").sum())
    n_analyzed = int(len(analysis_values))

    lower_p = (1.0 - cfg.central_fraction) / 2.0
    upper_p = 1.0 - lower_p

    if n_analyzed == 0:
        summary = {
            "partition": partition_name,
            "n_total": n_total,
            "n_analyzed": n_analyzed,
            "n_missing_or_non_numeric": n_missing,
            "n_manual_excluded": n_manual_excluded,
            "n_outlier_flagged": n_outlier_flagged,
            "lower_percentile": lower_p * 100,
            "upper_percentile": upper_p * 100,
            "lower_limit": np.nan,
            "upper_limit": np.nan,
            "lower_limit_CI_low": np.nan,
            "lower_limit_CI_high": np.nan,
            "upper_limit_CI_low": np.nan,
            "upper_limit_CI_high": np.nan,
            "lower_CI_rank_low": np.nan,
            "lower_CI_rank_high": np.nan,
            "upper_CI_rank_low": np.nan,
            "upper_CI_rank_high": np.nan,
            "method": "nonparametric percentile",
            "status": "No analyzable numeric values",
            "RI_report_text": "",
        }
        return summary, df

    lower_limit = percentile_nplus1(analysis_values, lower_p)
    upper_limit = percentile_nplus1(analysis_values, upper_p)

    lower_ci = order_stat_percentile_ci(analysis_values, lower_p, cfg.limit_ci)
    upper_ci = order_stat_percentile_ci(analysis_values, upper_p, cfg.limit_ci)

    if n_analyzed >= cfg.min_n_nonparametric:
        status = "OK: n>=120, nonparametric RI appropriate"
    else:
        status = (
            f"Caution: n<{cfg.min_n_nonparametric}. "
            "Nonparametric limits are shown, but EP28-A3c-style establishment is less stable; "
            "consider adding reference individuals or using this as preliminary/verification context."
        )

    unit_suffix = f" {cfg.units}" if cfg.units else ""
    ri_text = (
        f"{format_number(lower_limit, cfg.decimal_places)}"
        f"–{format_number(upper_limit, cfg.decimal_places)}{unit_suffix}"
    )
    lower_ci_text = (
        f"{format_number(lower_ci['ci_low'], cfg.decimal_places)}"
        f"–{format_number(lower_ci['ci_high'], cfg.decimal_places)}{unit_suffix}"
    )
    upper_ci_text = (
        f"{format_number(upper_ci['ci_low'], cfg.decimal_places)}"
        f"–{format_number(upper_ci['ci_high'], cfg.decimal_places)}{unit_suffix}"
    )

    summary = {
        "partition": partition_name,
        "n_total": n_total,
        "n_analyzed": n_analyzed,
        "n_missing_or_non_numeric": n_missing,
        "n_manual_excluded": n_manual_excluded,
        "n_outlier_flagged": n_outlier_flagged,
        "outliers_excluded_from_RI": bool(cfg.exclude_flagged_outliers),
        "lower_percentile": lower_p * 100,
        "upper_percentile": upper_p * 100,
        "lower_limit": lower_limit,
        "upper_limit": upper_limit,
        "lower_limit_CI_low": lower_ci["ci_low"],
        "lower_limit_CI_high": lower_ci["ci_high"],
        "upper_limit_CI_low": upper_ci["ci_low"],
        "upper_limit_CI_high": upper_ci["ci_high"],
        "lower_CI_rank_low": lower_ci["rank_low"],
        "lower_CI_rank_high": lower_ci["rank_high"],
        "upper_CI_rank_low": upper_ci["rank_low"],
        "upper_CI_rank_high": upper_ci["rank_high"],
        "lower_CI_actual_coverage": lower_ci["actual_coverage"],
        "upper_CI_actual_coverage": upper_ci["actual_coverage"],
        "method": "nonparametric percentile, p*(n+1) rank with interpolation",
        "limit_CI_method": "binomial/order-statistic CI",
        "status": status,
        "RI_report_text": ri_text,
        "lower_limit_90CI_text": lower_ci_text,
        "upper_limit_90CI_text": upper_ci_text,
    }
    return summary, df


def run_analysis(input_path: Path, sheet_name, cfg: AnalysisConfig, output_path: Path) -> None:
    df = pd.read_excel(input_path, sheet_name=sheet_name)

    # 입력 컬럼 검증
    required_cols = [cfg.value_col] + cfg.partition_cols
    if cfg.manual_exclude_col:
        required_cols.append(cfg.manual_exclude_col)
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(
            "엑셀 파일에서 다음 컬럼을 찾을 수 없습니다: " + ", ".join(missing_cols) +
            f"\n사용 가능한 컬럼: {list(df.columns)}"
        )

    # Partition key 생성
    if cfg.partition_cols:
        df["_partition"] = df.apply(lambda r: make_partition_key(r, cfg.partition_cols), axis=1)
    else:
        df["_partition"] = "ALL"

    summaries = []
    flagged_frames = []

    for partition_name, part_df in df.groupby("_partition", dropna=False, sort=True):
        summary, flagged = analyze_one_partition(part_df, cfg, str(partition_name))
        summaries.append(summary)
        flagged_frames.append(flagged)

    summary_df = pd.DataFrame(summaries)
    flagged_df = pd.concat(flagged_frames, axis=0).sort_index()
    harris_boyd_df = assess_harris_boyd_partitions(flagged_df, cfg)

    # 원본 컬럼 + 분석용 컬럼 순서 정리
    analysis_cols = [
        "_partition",
        "_value_numeric",
        "_missing_or_non_numeric",
        "_manual_exclude",
        "_outlier_flag",
    ]
    original_cols = [c for c in df.columns if c not in analysis_cols]
    flagged_df = flagged_df[original_cols + [c for c in analysis_cols if c in flagged_df.columns]]

    audit_rows = [
        ["analysis_datetime", datetime.now().isoformat(timespec="seconds")],
        ["input_file", str(input_path)],
        ["sheet_name", str(sheet_name)],
        ["value_col", cfg.value_col],
        ["partition_cols", ", ".join(cfg.partition_cols) if cfg.partition_cols else "ALL"],
        ["manual_exclude_col", cfg.manual_exclude_col or ""],
        ["central_fraction", cfg.central_fraction],
        ["reference_limits", f"{(1-cfg.central_fraction)/2*100:.3g}th and {(1+(cfg.central_fraction))/2*100:.3g}th percentiles"],
        ["limit_confidence_interval", cfg.limit_ci],
        ["minimum_n_for_nonparametric_establishment", cfg.min_n_nonparametric],
        ["outlier_method", cfg.outlier_method],
        ["exclude_flagged_outliers_from_RI", cfg.exclude_flagged_outliers],
        ["units", cfg.units],
        ["decimal_places", cfg.decimal_places],
        ["harris_boyd_enabled", cfg.harris_boyd_enabled],
        ["harris_boyd_transform", cfg.harris_boyd_transform],
        ["harris_boyd_min_n", cfg.harris_boyd_min_n],
        ["harris_boyd_z_formula", "Z=abs(mean1-mean2)/sqrt(sd1^2/n1 + sd2^2/n2)"],
        ["harris_boyd_zcritical_formula", "Z*=3*sqrt((n1+n2)/240)"],
        ["harris_boyd_primary_decision_rule", "Recommend partitioning when Z > Z*"],
        ["harris_boyd_auxiliary_sd_flags", "Review when larger SD/smaller SD >=1.5 or s_large/(s_large-s_small)<3"],
        ["important_note", "Outlier flags are for professional review. Do not exclude values solely by statistical flag without documented analytical/preanalytical/clinical reason."],
        ["harris_boyd_note", "Harris-Boyd assumes approximately Gaussian subgroup distributions. Use histograms/Q-Q plots and clinical judgment before final partition decisions."],
    ]
    audit_df = pd.DataFrame(audit_rows, columns=["item", "value"])

    # 엑셀 출력
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="RI_summary")
        harris_boyd_df.to_excel(writer, index=False, sheet_name="Partition_HB")
        flagged_df.to_excel(writer, index=False, sheet_name="Data_with_flags")
        audit_df.to_excel(writer, index=False, sheet_name="Audit")

    style_workbook(output_path)

    print(f"\n완료: {output_path}")
    print("\n요약:")
    display_cols = ["partition", "n_analyzed", "RI_report_text", "lower_limit_90CI_text", "upper_limit_90CI_text", "status"]
    existing_display_cols = [c for c in display_cols if c in summary_df.columns]
    print(summary_df[existing_display_cols].to_string(index=False))

    if not harris_boyd_df.empty:
        print("\nHarris-Boyd partition assessment:")
        hb_display_cols = [
            "partition_1", "partition_2", "n1", "n2",
            "z_calculated", "z_critical", "z_gt_zcritical",
            "sd_ratio_large_to_small", "decision"
        ]
        existing_hb_cols = [c for c in hb_display_cols if c in harris_boyd_df.columns]
        print(harris_boyd_df[existing_hb_cols].to_string(index=False))


# -----------------------------
# Excel styling
# -----------------------------

def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin_gray = Side(style="thin", color="BFBFBF")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        # column width 조정
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    wb.save(path)


# -----------------------------
# CLI
# -----------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="CLSI EP28-A3c 스타일 참고구간 분석: nonparametric percentile + order-statistic CI"
    )
    parser.add_argument("input", help="입력 엑셀 파일 경로, 예: data.xlsx")
    parser.add_argument("--sheet", default="0", help="시트명 또는 0-based 시트 번호. 기본값: 0")
    parser.add_argument("--value-col", required=True, help="검사결과 값 컬럼명")
    parser.add_argument(
        "--partition-cols",
        nargs="*",
        default=[],
        help="partition 컬럼명들. 예: --partition-cols Sex AgeGroup. 생략하면 전체 분석",
    )
    parser.add_argument(
        "--manual-exclude-col",
        default=None,
        help="수동 제외 여부 컬럼명. 1/TRUE/yes/제외 등을 제외로 처리",
    )
    parser.add_argument(
        "--central-fraction",
        type=float,
        default=0.95,
        help="중앙 참고구간 비율. 기본값 0.95 -> 2.5~97.5 percentile",
    )
    parser.add_argument(
        "--limit-ci",
        type=float,
        default=0.90,
        help="각 reference limit의 confidence interval. 기본값 0.90",
    )
    parser.add_argument(
        "--min-n-nonparametric",
        type=int,
        default=120,
        help="비모수 RI 설정에 권장되는 최소 n. 기본값 120",
    )
    parser.add_argument(
        "--outlier-method",
        choices=["none", "tukey", "reed", "both"],
        default="both",
        help="이상치 flag 방식. 기본값 both",
    )
    parser.add_argument(
        "--exclude-flagged-outliers",
        action="store_true",
        help="flag된 이상치를 RI 계산에서 제외. 기본값은 제외하지 않음. 검토 후에만 사용 권장",
    )
    parser.add_argument("--units", default="", help="보고서에 표시할 단위, 예: U/L")
    parser.add_argument("--decimal-places", type=int, default=3, help="출력 소수점 자리수. 기본값 3")
    parser.add_argument(
        "--skip-harris-boyd",
        action="store_true",
        help="partition 컬럼이 있어도 Harris-Boyd partition 평가를 수행하지 않음",
    )
    parser.add_argument(
        "--hb-transform",
        choices=["none", "log", "boxcox"],
        default="none",
        help="Harris-Boyd 계산 전 변환 방식. 기본값 none. log/boxcox는 모든 값이 >0일 때만 적용",
    )
    parser.add_argument(
        "--hb-min-n",
        type=int,
        default=3,
        help="Harris-Boyd pairwise 비교에 필요한 각 partition의 최소 숫자값 개수. 기본값 3",
    )
    parser.add_argument(
        "--normality-alpha",
        type=float,
        default=0.05,
        help="Harris-Boyd 보조 정규성 검토에서 사용할 Shapiro-Wilk alpha. 기본값 0.05",
    )
    parser.add_argument("--out", default="ep28_RI_report.xlsx", help="출력 엑셀 파일명")
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    if not (0 < args.central_fraction < 1):
        raise ValueError("--central-fraction은 0과 1 사이여야 합니다.")
    if not (0 < args.limit_ci < 1):
        raise ValueError("--limit-ci는 0과 1 사이여야 합니다.")
    if args.decimal_places < 0:
        raise ValueError("--decimal-places는 0 이상이어야 합니다.")
    if args.hb_min_n < 2:
        raise ValueError("--hb-min-n은 2 이상이어야 합니다.")
    if not (0 < args.normality_alpha < 1):
        raise ValueError("--normality-alpha는 0과 1 사이여야 합니다.")

    cfg = AnalysisConfig(
        value_col=args.value_col,
        partition_cols=args.partition_cols,
        manual_exclude_col=args.manual_exclude_col,
        central_fraction=args.central_fraction,
        limit_ci=args.limit_ci,
        min_n_nonparametric=args.min_n_nonparametric,
        outlier_method=args.outlier_method,
        exclude_flagged_outliers=args.exclude_flagged_outliers,
        units=args.units,
        decimal_places=args.decimal_places,
        harris_boyd_enabled=not args.skip_harris_boyd,
        harris_boyd_transform=args.hb_transform,
        harris_boyd_min_n=args.hb_min_n,
        normality_alpha=args.normality_alpha,
    )

    run_analysis(
        input_path=Path(args.input),
        sheet_name=parse_sheet_arg(str(args.sheet)),
        cfg=cfg,
        output_path=Path(args.out),
    )


if __name__ == "__main__":
    main()
